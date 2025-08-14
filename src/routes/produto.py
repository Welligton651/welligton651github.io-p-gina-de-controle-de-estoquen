from flask import Blueprint, request, jsonify
from src.models.produto import db, Produto, MovimentacaoEstoque
from sqlalchemy import or_
import pandas as pd
import io
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

produto_bp = Blueprint('produto', __name__)

@produto_bp.route('/api/produtos', methods=['GET'])
def listar_produtos():
    """Lista todos os produtos com filtros opcionais"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '')
        status = request.args.get('status', '')
        
        query = Produto.query
        
        # Filtro de busca por descrição
        if search:
            query = query.filter(Produto.descricao.contains(search))
        
        # Filtro por status do estoque
        if status == 'ESGOTADO':
            query = query.filter(Produto.estoque <= 0)
        elif status == 'BAIXO':
            query = query.filter(Produto.estoque <= Produto.estoque_minimo, Produto.estoque > 0)
        elif status == 'OK':
            query = query.filter(Produto.estoque > Produto.estoque_minimo)
        
        # Ordenação por descrição
        query = query.order_by(Produto.descricao)
        
        # Paginação
        produtos = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        return jsonify({
            'produtos': [produto.to_dict() for produto in produtos.items],
            'total': produtos.total,
            'pages': produtos.pages,
            'current_page': page,
            'per_page': per_page
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@produto_bp.route('/api/produtos/<int:produto_id>', methods=['GET'])
def obter_produto(produto_id):
    """Obtém um produto específico"""
    try:
        produto = Produto.query.get_or_404(produto_id)
        return jsonify(produto.to_dict())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@produto_bp.route('/api/produtos', methods=['POST'])
def criar_produto():
    """Cria um novo produto"""
    try:
        data = request.get_json()
        
        if not data or not data.get('descricao'):
            return jsonify({'error': 'Descrição é obrigatória'}), 400
        
        produto = Produto(
            descricao=data['descricao'],
            unidade=data.get('unidade', 'UNIDADE'),
            fornecimento=float(data.get('fornecimento', 0)),
            estoque=float(data.get('estoque', 0)),
            estoque_minimo=float(data.get('estoque_minimo', 5))
        )
        
        db.session.add(produto)
        db.session.commit()
        
        return jsonify(produto.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@produto_bp.route('/api/produtos/<int:produto_id>', methods=['PUT'])
def atualizar_produto(produto_id):
    """Atualiza um produto existente"""
    try:
        produto = Produto.query.get_or_404(produto_id)
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Dados não fornecidos'}), 400
        
        produto.descricao = data.get('descricao', produto.descricao)
        produto.unidade = data.get('unidade', produto.unidade)
        produto.fornecimento = float(data.get('fornecimento', produto.fornecimento))
        produto.estoque = float(data.get('estoque', produto.estoque))
        produto.estoque_minimo = float(data.get('estoque_minimo', produto.estoque_minimo))
        produto.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify(produto.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@produto_bp.route('/api/produtos/<int:produto_id>', methods=['DELETE'])
def deletar_produto(produto_id):
    """Deleta um produto"""
    try:
        produto = Produto.query.get_or_404(produto_id)
        db.session.delete(produto)
        db.session.commit()
        
        return jsonify({'message': 'Produto deletado com sucesso'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@produto_bp.route('/api/produtos/<int:produto_id>/baixa', methods=['POST'])
def dar_baixa_estoque(produto_id):
    """Dá baixa no estoque de um produto"""
    try:
        produto = Produto.query.get_or_404(produto_id)
        data = request.get_json()
        
        if not data or 'quantidade' not in data:
            return jsonify({'error': 'Quantidade é obrigatória'}), 400
        
        quantidade = float(data['quantidade'])
        observacao = data.get('observacao', '')
        
        if quantidade <= 0:
            return jsonify({'error': 'Quantidade deve ser maior que zero'}), 400
        
        if produto.estoque < quantidade:
            return jsonify({'error': f'Estoque insuficiente. Disponível: {produto.estoque}'}), 400
        
        # Dar baixa no estoque
        produto.dar_baixa(quantidade)
        
        # Registrar movimentação
        movimentacao = MovimentacaoEstoque(
            produto_id=produto.id,
            tipo='SAIDA',
            quantidade=quantidade,
            observacao=observacao
        )
        
        db.session.add(movimentacao)
        db.session.commit()
        
        return jsonify({
            'message': 'Baixa realizada com sucesso',
            'produto': produto.to_dict(),
            'movimentacao': movimentacao.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@produto_bp.route('/api/produtos/<int:produto_id>/entrada', methods=['POST'])
def adicionar_estoque(produto_id):
    """Adiciona quantidade ao estoque de um produto"""
    try:
        produto = Produto.query.get_or_404(produto_id)
        data = request.get_json()
        
        if not data or 'quantidade' not in data:
            return jsonify({'error': 'Quantidade é obrigatória'}), 400
        
        quantidade = float(data['quantidade'])
        observacao = data.get('observacao', '')
        
        if quantidade <= 0:
            return jsonify({'error': 'Quantidade deve ser maior que zero'}), 400
        
        # Adicionar ao estoque
        produto.adicionar_estoque(quantidade)
        
        # Registrar movimentação
        movimentacao = MovimentacaoEstoque(
            produto_id=produto.id,
            tipo='ENTRADA',
            quantidade=quantidade,
            observacao=observacao
        )
        
        db.session.add(movimentacao)
        db.session.commit()
        
        return jsonify({
            'message': 'Entrada realizada com sucesso',
            'produto': produto.to_dict(),
            'movimentacao': movimentacao.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@produto_bp.route('/api/movimentacoes', methods=['GET'])
def listar_movimentacoes():
    """Lista as movimentações de estoque"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        produto_id = request.args.get('produto_id', type=int)
        tipo = request.args.get('tipo', '')
        
        query = MovimentacaoEstoque.query
        
        if produto_id:
            query = query.filter(MovimentacaoEstoque.produto_id == produto_id)
        
        if tipo:
            query = query.filter(MovimentacaoEstoque.tipo == tipo)
        
        query = query.order_by(MovimentacaoEstoque.data_movimentacao.desc())
        
        movimentacoes = query.paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        return jsonify({
            'movimentacoes': [mov.to_dict() for mov in movimentacoes.items],
            'total': movimentacoes.total,
            'pages': movimentacoes.pages,
            'current_page': page,
            'per_page': per_page
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@produto_bp.route('/api/dashboard', methods=['GET'])
def dashboard():
    """Retorna dados para o dashboard"""
    try:
        total_produtos = Produto.query.count()
        produtos_esgotados = Produto.query.filter(Produto.estoque <= 0).count()
        produtos_baixo_estoque = Produto.query.filter(
            Produto.estoque <= Produto.estoque_minimo,
            Produto.estoque > 0
        ).count()
        
        # Produtos com estoque crítico
        produtos_criticos = Produto.query.filter(
            Produto.estoque <= Produto.estoque_minimo
        ).order_by(Produto.estoque).limit(10).all()
        
        # Últimas movimentações
        ultimas_movimentacoes = MovimentacaoEstoque.query.order_by(
            MovimentacaoEstoque.data_movimentacao.desc()
        ).limit(10).all()
        
        return jsonify({
            'total_produtos': total_produtos,
            'produtos_esgotados': produtos_esgotados,
            'produtos_baixo_estoque': produtos_baixo_estoque,
            'produtos_ok': total_produtos - produtos_esgotados - produtos_baixo_estoque,
            'produtos_criticos': [p.to_dict() for p in produtos_criticos],
            'ultimas_movimentacoes': [m.to_dict() for m in ultimas_movimentacoes]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@produto_bp.route('/api/exportar-xlsx', methods=['GET'])
def exportar_xlsx():
    """Exporta todos os produtos para um arquivo XLSX"""
    try:
        produtos = Produto.query.order_by(Produto.descricao).all()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Controle de Estoque"
        
        # Cabeçalhos
        headers = ['DESCRIÇÃO DO ITEM', 'UNIDADE', 'FORNECIMENTO', 'ESTOQUE']
        ws.append(headers)
        
        # Estilizar cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Adicionar dados
        for produto in produtos:
            ws.append([
                produto.descricao,
                produto.unidade or 'UNIDADE',
                produto.fornecimento,
                produto.estoque
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Salvar em buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'controle_estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

